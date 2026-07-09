from flask import Flask, jsonify, request, send_from_directory, session, redirect, url_for, render_template
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import timedelta

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = 'investify-session-secret-key-2026'
app.permanent_session_lifetime = timedelta(days=7)

JSON_PATH = os.path.join(os.path.dirname(__file__), 'db.json')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Investimentos.xlsx')
EXPORT_PATH = os.path.join(os.path.dirname(__file__), 'Investimentos_Exportado.xlsx')
CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config.json')

def load_credentials():
    if os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {"username": "admin", "password": "change-me-on-first-login"}

def load_db():
    if not os.path.exists(JSON_PATH):
        # Try to parse from Excel if JSON doesn't exist
        try:
            import parse_data
            parse_data.main()
        except Exception as e:
            print("Error parsing Excel:", e)
            
    try:
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_db(data):
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

@app.before_request
def check_authentication():
    allowed_routes = ['login', 'do_login', 'static']
    if request.endpoint in allowed_routes or not request.endpoint:
        return
    
    if 'logged_in' not in session:
        if request.path.startswith('/api/'):
            return jsonify({"error": "Sessão expirada ou não autenticada."}), 401
        return redirect(url_for('login'))

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login():
    if 'logged_in' in session:
        return redirect(url_for('index'))
    return send_from_directory('static', 'login.html')

@app.route('/api/login', methods=['POST'])
def do_login():
    payload = request.json or {}
    username = payload.get('username')
    password = payload.get('password')
    
    creds = load_credentials()
    if username == creds.get('username') and password == creds.get('password'):
        session['logged_in'] = True
        session.permanent = True
        return jsonify({"success": True})
    
    return jsonify({"error": "Usuário ou senha incorretos."}), 401

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/api/data', methods=['GET'])
def get_data():
    db = load_db()
    return jsonify(db)

@app.route('/api/entry', methods=['POST'])
def save_entry():
    payload = request.json
    year = str(payload.get('year'))
    month = str(payload.get('month'))
    
    if not year or not month:
        return jsonify({"error": "Ano e mês são obrigatórios"}), 400
        
    db = load_db()
    
    if year not in db:
        db[year] = {}
        
    # Standardize data model
    investments = payload.get('investments', {})
    expenses = payload.get('expenses', {})
    incomes = payload.get('incomes', {})
    
    # Calculate Juros automatically if not provided or set to 0 and we have a previous month
    # Formula: Juros(M) = CDB(M) - (CDB(M-1) + Aporte(M-1) - Retirada(M-1))
    prev_month = str(int(month) - 1)
    for bank in ['itau', 'bb', 'c6']:
        bank_data = investments.get(bank, {})
        cdb_val = float(bank_data.get('cdb', 0.0))
        ap_val = float(bank_data.get('aporte', 0.0))
        ret_val = float(bank_data.get('retirada', 0.0))
        jr_val = bank_data.get('juros')
        
        # If juros is not provided explicitly or is empty, we calculate it
        if jr_val is None or jr_val == "":
            calculated_juros = 0.0
            if prev_month in db[year]:
                prev_bank_data = db[year][prev_month]['investments'].get(bank, {})
                prev_cdb = float(prev_bank_data.get('cdb', 0.0))
                prev_ap = float(prev_bank_data.get('aporte', 0.0))
                prev_ret = float(prev_bank_data.get('retirada', 0.0))
                calculated_juros = cdb_val - (prev_cdb + prev_ap - prev_ret)
            elif int(month) == 1:
                # If January, look for December of previous year
                prev_year = str(int(year) - 1)
                if prev_year in db and '12' in db[prev_year]:
                    prev_bank_data = db[prev_year]['12']['investments'].get(bank, {})
                    prev_cdb = float(prev_bank_data.get('cdb', 0.0))
                    prev_ap = float(prev_bank_data.get('aporte', 0.0))
                    prev_ret = float(prev_bank_data.get('retirada', 0.0))
                    calculated_juros = cdb_val - (prev_cdb + prev_ap - prev_ret)
            bank_data['juros'] = round(calculated_juros, 2)
        else:
            bank_data['juros'] = round(float(jr_val), 2)
            
        # Calculate bank total (CDB + Aporte + Juros - Retirada)
        bank_tot = cdb_val + ap_val + float(bank_data.get('juros', 0.0)) - ret_val
        bank_data['total'] = round(bank_tot, 2)
        bank_data['cdb'] = round(cdb_val, 2)
        bank_data['aporte'] = round(ap_val, 2)
        bank_data['retirada'] = round(ret_val, 2)
        
    portfolio_tot = sum(float(investments[b].get('total', 0.0)) for b in ['itau', 'bb', 'c6'])
        
    db[year][month] = {
        "investments": investments,
        "total_carteira": round(portfolio_tot, 2),
        "expenses": {
            "fixed": {k: round(float(v), 2) for k, v in expenses.get('fixed', {}).items()},
            "cards": {k: round(float(v), 2) for k, v in expenses.get('cards', {}).items()}
        },
        "incomes": {k: round(float(v), 2) for k, v in incomes.items()}
    }
    
    save_db(db)
    return jsonify({"success": True, "message": "Lançamento salvo com sucesso", "data": db[year][month]})

@app.route('/api/import-gsheet', methods=['POST'])
def import_gsheet():
    import urllib.request
    import urllib.error
    import re
    
    payload = request.json
    url = payload.get('url')
    if not url:
        return jsonify({"error": "URL do Google Sheets é obrigatória"}), 400
        
    # Extract spreadsheet ID from url
    match = re.search(r'/d/([a-zA-Z0-9-_]+)', url)
    if not match:
        return jsonify({"error": "Link do Google Sheets inválido. Certifique-se de que é um link válido de compartilhamento."}), 400
        
    spreadsheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    
    temp_path = os.path.join(os.path.dirname(__file__), 'Investimentos_Online.xlsx')
    
    try:
        # Download spreadsheet
        req = urllib.request.Request(
            export_url, 
            headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)'}
        )
        with urllib.request.urlopen(req) as response:
            with open(temp_path, 'wb') as out_file:
                out_file.write(response.read())
                
        # Parse downloaded file using parse_data
        import parse_data
        parse_data.main(temp_path)
        
        # Replace the local spreadsheet file so it is synced!
        if os.path.exists(EXCEL_PATH):
            backup_path = EXCEL_PATH + ".bak"
            if os.path.exists(backup_path):
                os.remove(backup_path)
            os.rename(EXCEL_PATH, backup_path)
        os.rename(temp_path, EXCEL_PATH)
        
        return jsonify({"success": True, "message": "Planilha online importada e integrada com sucesso!"})
        
    except urllib.error.HTTPError as e:
        if e.code == 401 or e.code == 403:
            return jsonify({
                "error": "Acesso não autorizado (HTTP 401/403). Por favor, altere o acesso da sua planilha online para 'Qualquer pessoa com o link pode ler' nas configurações de compartilhamento do Google Drive."
            }), 403
        return jsonify({"error": f"Erro de download do Google Sheets: {e.reason} (HTTP {e.code})"}), 500
    except Exception as e:
        return jsonify({"error": f"Erro ao processar planilha: {str(e)}"}), 500

@app.route('/api/export', methods=['POST'])
def export_excel():
    """
    Generates a professionally formatted Excel workbook from db.json.
    """
    db = load_db()
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    section_font = Font(name=font_family, size=12, bold=True, color="1F497D")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True)
    regular_font = Font(name=font_family, size=10)
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_accent = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    fill_total = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=Side(border_style="thin", color="1F497D"), bottom=Side(border_style="double", color="1F497D"))
    
    months_names = [
        "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO", 
        "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
    ]
    
    for year in sorted(db.keys(), reverse=True):
        ws = wb.create_sheet(title=str(year))
        ws.views.sheetView[0].showGridLines = True
        
        # Title
        ws['A1'] = f"Controle Financeiro e de Investimentos - {year}"
        ws['A1'].font = title_font
        
        # Table 1: Investments
        ws['A3'] = "INVESTIMENTOS"
        ws['A3'].font = section_font
        
        # Headers for Investments
        ws.cell(row=4, column=1, value="Instituição / Tipo").font = header_font
        ws.cell(row=4, column=1).fill = fill_header
        ws.cell(row=4, column=2, value="Métrica").font = header_font
        ws.cell(row=4, column=2).fill = fill_header
        
        for m_idx, m_name in enumerate(months_names):
            c_cell = ws.cell(row=4, column=3 + m_idx, value=m_name)
            c_cell.font = header_font
            c_cell.fill = fill_header
            c_cell.alignment = Alignment(horizontal="center")
            
        ws.cell(row=4, column=15, value="TOTAL ANO").font = header_font
        ws.cell(row=4, column=15).fill = fill_header
        
        # Fill data row by row
        # We have banks: Itaú, BB, C6
        banks_mapping = [
            ("ITAÚ", "itau"),
            ("BANCO DO BRASIL", "bb"),
            ("C6 BANK", "c6")
        ]
        
        row_cursor = 5
        for bank_label, bank_key in banks_mapping:
            metrics = [
                ("CDB, Renda Fixa", "cdb"),
                ("Juros / Rendimento", "juros"),
                ("Aporte", "aporte"),
                ("Retirada", "retirada"),
                ("Total", "total")
            ]
            
            # Write bank title
            ws.cell(row=row_cursor, column=1, value=bank_label).font = bold_font
            ws.cell(row=row_cursor, column=1).fill = fill_accent
            
            for m_label, m_key in metrics:
                ws.cell(row=row_cursor, column=2, value=m_label).font = bold_font if m_key == "total" else regular_font
                
                # Monthly values
                for m in range(1, 13):
                    val = 0.0
                    month_data = db[year].get(str(m), {})
                    bank_data = month_data.get("investments", {}).get(bank_key, {})
                    
                    if m_key == "total":
                        val = float(bank_data.get("cdb", 0.0)) + float(bank_data.get("aporte", 0.0)) + float(bank_data.get("juros", 0.0)) - float(bank_data.get("retirada", 0.0))
                    else:
                        val = float(bank_data.get(m_key, 0.0))
                        
                    cell = ws.cell(row=row_cursor, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font if m_key == "total" else regular_font
                    cell.border = border_all
                    
                # Annual total formula
                col_start = get_column_letter(3)
                col_end = get_column_letter(14)
                
                if m_key == "cdb" or m_key == "total":
                    # CDB or Total is balance, so we show the last non-zero or December balance
                    cell_tot = ws.cell(row=row_cursor, column=15, value=f"=O{row_cursor}")
                else:
                    cell_tot = ws.cell(row=row_cursor, column=15, value=f"=SUM({col_start}{row_cursor}:{col_end}{row_cursor})")
                    
                cell_tot.number_format = '#,##0.00'
                cell_tot.font = bold_font
                cell_tot.border = border_all
                
                row_cursor += 1
            row_cursor += 1 # Empty line
            
        # Consolidation Row (JUROS MENSAIS, APORTES, RETIRADAS, TOTAL DO MÊS)
        ws.cell(row=row_cursor, column=1, value="CONSOLIDAÇÃO INVESTIMENTOS").font = bold_font
        row_cursor += 1
        
        consol_metrics = [
            ("JUROS MENSAIS", "juros"),
            ("APORTES MENSAIS", "aporte"),
            ("RETIRADAS MENSAIS", "retirada"),
            ("VALOR TOTAL CARTEIRA", "total")
        ]
        
        for c_label, c_key in consol_metrics:
            ws.cell(row=row_cursor, column=2, value=c_label).font = bold_font
            
            for m in range(1, 13):
                col_letter = get_column_letter(3 + m - 1)
                
                if c_key == "juros":
                    # Sum rows: Itaú Juros (Row 6), BB Juros (Row 12), C6 Juros (Row 18)
                    formula = f"={col_letter}6+{col_letter}12+{col_letter}18"
                elif c_key == "aporte":
                    # Sum rows: Itaú Aporte (Row 7), BB Aporte (Row 13), C6 Aporte (Row 19)
                    formula = f"={col_letter}7+{col_letter}13+{col_letter}19"
                elif c_key == "retirada":
                    # Sum rows: Itaú Retirada (Row 8), BB Retirada (Row 14), C6 Retirada (Row 20)
                    formula = f"={col_letter}8+{col_letter}14+{col_letter}20"
                else:
                    # Sum rows: Itaú Total (Row 9), BB Total (Row 15), C6 Total (Row 21)
                    formula = f"={col_letter}9+{col_letter}15+{col_letter}21"
                    
                cell = ws.cell(row=row_cursor, column=3 + m - 1, value=formula)
                cell.number_format = '#,##0.00'
                cell.font = bold_font
                cell.border = border_all
                if c_key == "total":
                    cell.fill = fill_total
                    
            # Formula for Consolidation Total Year
            if c_key == "total":
                cell_tot = ws.cell(row=row_cursor, column=15, value="=O" + str(row_cursor))
            else:
                cell_tot = ws.cell(row=row_cursor, column=15, value=f"=SUM(C{row_cursor}:N{row_cursor})")
                
            cell_tot.number_format = '#,##0.00'
            cell_tot.font = bold_font
            cell_tot.border = border_all
            if c_key == "total":
                cell_tot.fill = fill_total
                
            row_cursor += 1
            
        row_cursor += 2 # Empty spacing
        
        # Table 2: Budget (Fixed Expenses, Cards, Incomes)
        # Check if year has expenses
        sample_month = list(db[year].values())[0] if db[year] else {}
        fixed_exp_keys = list(sample_month.get("expenses", {}).get("fixed", {}).keys())
        card_keys = list(sample_month.get("expenses", {}).get("cards", {}).keys())
        income_keys = list(sample_month.get("incomes", {}).keys())
        
        if fixed_exp_keys:
            ws.cell(row=row_cursor, column=1, value="ORÇAMENTO E FLUXO DE CAIXA").font = section_font
            row_cursor += 1
            
            # Gastos Fixos Header
            ws.cell(row=row_cursor, column=1, value="GASTOS FIXOS").font = bold_font
            row_cursor += 1
            
            fixed_start_row = row_cursor
            for exp_key in fixed_exp_keys:
                ws.cell(row=row_cursor, column=2, value=exp_key.upper().replace('_', ' ')).font = regular_font
                for m in range(1, 13):
                    val = float(db[year].get(str(m), {}).get("expenses", {}).get("fixed", {}).get(exp_key, 0.0))
                    cell = ws.cell(row=row_cursor, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = regular_font
                    cell.border = border_all
                # Sum Row
                cell_tot = ws.cell(row=row_cursor, column=15, value=f"=SUM(C{row_cursor}:N{row_cursor})")
                cell_tot.number_format = '#,##0.00'
                cell_tot.font = bold_font
                cell_tot.border = border_all
                row_cursor += 1
                
            # Gastos Fixos Total
            fixed_end_row = row_cursor - 1
            ws.cell(row=row_cursor, column=2, value="TOTAL GASTOS FIXOS").font = bold_font
            for m in range(1, 13):
                col_letter = get_column_letter(3 + m - 1)
                cell = ws.cell(row=row_cursor, column=3 + m - 1, value=f"=SUM({col_letter}{fixed_start_row}:{col_letter}{fixed_end_row})")
                cell.number_format = '#,##0.00'
                cell.font = bold_font
                cell.border = border_total
            ws.cell(row=row_cursor, column=15, value=f"=SUM(O{fixed_start_row}:O{fixed_end_row})").font = bold_font
            ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
            ws.cell(row=row_cursor, column=15).border = border_total
            fixed_tot_row = row_cursor
            row_cursor += 2
            
            # Credit Cards Section (if any)
            cards_tot_row = None
            if card_keys:
                ws.cell(row=row_cursor, column=1, value="CARTÕES DE CRÉDITO").font = bold_font
                row_cursor += 1
                
                cards_start_row = row_cursor
                for card_key in card_keys:
                    ws.cell(row=row_cursor, column=2, value=card_key.upper().replace('_', ' ')).font = regular_font
                    for m in range(1, 13):
                        val = float(db[year].get(str(m), {}).get("expenses", {}).get("cards", {}).get(card_key, 0.0))
                        cell = ws.cell(row=row_cursor, column=3 + m - 1, value=val)
                        cell.number_format = '#,##0.00'
                        cell.font = regular_font
                        cell.border = border_all
                    # Sum Row
                    cell_tot = ws.cell(row=row_cursor, column=15, value=f"=SUM(C{row_cursor}:N{row_cursor})")
                    cell_tot.number_format = '#,##0.00'
                    cell_tot.font = bold_font
                    cell_tot.border = border_all
                    row_cursor += 1
                    
                cards_end_row = row_cursor - 1
                ws.cell(row=row_cursor, column=2, value="TOTAL CARTÕES").font = bold_font
                for m in range(1, 13):
                    col_letter = get_column_letter(3 + m - 1)
                    cell = ws.cell(row=row_cursor, column=3 + m - 1, value=f"=SUM({col_letter}{cards_start_row}:{col_letter}{cards_end_row})")
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font
                    cell.border = border_total
                ws.cell(row=row_cursor, column=15, value=f"=SUM(O{cards_start_row}:O{cards_end_row})").font = bold_font
                ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
                ws.cell(row=row_cursor, column=15).border = border_total
                cards_tot_row = row_cursor
                row_cursor += 2
                
            # Total Expenses Consolidation
            ws.cell(row=row_cursor, column=1, value="DESPESAS TOTAIS").font = bold_font
            for m in range(1, 13):
                col_letter = get_column_letter(3 + m - 1)
                formula = f"={col_letter}{fixed_tot_row}"
                if cards_tot_row:
                    formula += f"+{col_letter}{cards_tot_row}"
                cell = ws.cell(row=row_cursor, column=3 + m - 1, value=formula)
                cell.number_format = '#,##0.00'
                cell.font = bold_font
                cell.border = border_total
                cell.fill = fill_accent
            formula_tot = f"=SUM(C{row_cursor}:N{row_cursor})"
            ws.cell(row=row_cursor, column=15, value=formula_tot).font = bold_font
            ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
            ws.cell(row=row_cursor, column=15).border = border_total
            ws.cell(row=row_cursor, column=15).fill = fill_accent
            expenses_tot_row = row_cursor
            row_cursor += 2
            
            # Incomes Section
            ws.cell(row=row_cursor, column=1, value="RECEITAS").font = bold_font
            row_cursor += 1
            
            incomes_start_row = row_cursor
            if str(year) == '2026':
                r15 = row_cursor
                ws.cell(row=r15, column=2, value="SALÁRIO DIA 15").font = regular_font
                for m in range(1, 13):
                    val = float(db[year].get(str(m), {}).get("incomes", {}).get("salario_dia_15", 0.0))
                    cell = ws.cell(row=r15, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = regular_font
                    cell.border = border_all
                ws.cell(row=r15, column=15, value=f"=SUM(C{r15}:N{r15})").font = bold_font
                ws.cell(row=r15, column=15).number_format = '#,##0.00'
                ws.cell(row=r15, column=15).border = border_all
                
                r30 = row_cursor + 1
                ws.cell(row=r30, column=2, value="SALÁRIO DIA 30").font = regular_font
                for m in range(1, 13):
                    val = float(db[year].get(str(m), {}).get("incomes", {}).get("salario_dia_30", 0.0))
                    cell = ws.cell(row=r30, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = regular_font
                    cell.border = border_all
                ws.cell(row=r30, column=15, value=f"=SUM(C{r30}:N{r30})").font = bold_font
                ws.cell(row=r30, column=15).number_format = '#,##0.00'
                ws.cell(row=r30, column=15).border = border_all
                
                rsum = row_cursor + 2
                ws.cell(row=rsum, column=2, value="SOMATÓRIO SALÁRIO").font = bold_font
                for m in range(1, 13):
                    col_letter = get_column_letter(3 + m - 1)
                    cell = ws.cell(row=rsum, column=3 + m - 1, value=f"=SUM({col_letter}{r15}:{col_letter}{r30})")
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font
                    cell.border = border_all
                ws.cell(row=rsum, column=15, value=f"=SUM(O{r15}:O{r30})").font = bold_font
                ws.cell(row=rsum, column=15).number_format = '#,##0.00'
                ws.cell(row=rsum, column=15).border = border_all
                
                raluguel = row_cursor + 3
                ws.cell(row=raluguel, column=2, value="ALUGUEL RECEBIDO").font = regular_font
                for m in range(1, 13):
                    val = float(db[year].get(str(m), {}).get("incomes", {}).get("aluguel", 0.0))
                    cell = ws.cell(row=raluguel, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = regular_font
                    cell.border = border_all
                ws.cell(row=raluguel, column=15, value=f"=SUM(C{raluguel}:N{raluguel})").font = bold_font
                ws.cell(row=raluguel, column=15).number_format = '#,##0.00'
                ws.cell(row=raluguel, column=15).border = border_all
                
                rextra = row_cursor + 4
                ws.cell(row=rextra, column=2, value="EXTRA/FÉRIAS/PLR").font = regular_font
                for m in range(1, 13):
                    val = float(db[year].get(str(m), {}).get("incomes", {}).get("extra", 0.0))
                    cell = ws.cell(row=rextra, column=3 + m - 1, value=val)
                    cell.number_format = '#,##0.00'
                    cell.font = regular_font
                    cell.border = border_all
                ws.cell(row=rextra, column=15, value=f"=SUM(C{rextra}:N{rextra})").font = bold_font
                ws.cell(row=rextra, column=15).number_format = '#,##0.00'
                ws.cell(row=rextra, column=15).border = border_all
                
                row_cursor += 5
                
                ws.cell(row=row_cursor, column=2, value="TOTAL RECEITAS").font = bold_font
                for m in range(1, 13):
                    col_letter = get_column_letter(3 + m - 1)
                    cell = ws.cell(row=row_cursor, column=3 + m - 1, value=f"={col_letter}{rsum}+{col_letter}{raluguel}+{col_letter}{rextra}")
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font
                    cell.border = border_total
                    cell.fill = fill_accent
                ws.cell(row=row_cursor, column=15, value=f"=SUM(O{rsum},O{raluguel},O{rextra})").font = bold_font
                ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
                ws.cell(row=row_cursor, column=15).border = border_total
                ws.cell(row=row_cursor, column=15).fill = fill_accent
                incomes_tot_row = row_cursor
                row_cursor += 2
            else:
                for inc_key in income_keys:
                    ws.cell(row=row_cursor, column=2, value=inc_key.upper().replace('_', ' ')).font = regular_font
                    for m in range(1, 13):
                        val = float(db[year].get(str(m), {}).get("incomes", {}).get(inc_key, 0.0))
                        cell = ws.cell(row=row_cursor, column=3 + m - 1, value=val)
                        cell.number_format = '#,##0.00'
                        cell.font = regular_font
                        cell.border = border_all
                    cell_tot = ws.cell(row=row_cursor, column=15, value=f"=SUM(C{row_cursor}:N{row_cursor})")
                    cell_tot.number_format = '#,##0.00'
                    cell_tot.font = bold_font
                    cell_tot.border = border_all
                    row_cursor += 1
                    
                incomes_end_row = row_cursor - 1
                ws.cell(row=row_cursor, column=2, value="TOTAL RECEITAS").font = bold_font
                for m in range(1, 13):
                    col_letter = get_column_letter(3 + m - 1)
                    cell = ws.cell(row=row_cursor, column=3 + m - 1, value=f"=SUM({col_letter}{incomes_start_row}:{col_letter}{incomes_end_row})")
                    cell.number_format = '#,##0.00'
                    cell.font = bold_font
                    cell.border = border_total
                    cell.fill = fill_accent
                ws.cell(row=row_cursor, column=15, value=f"=SUM(O{incomes_start_row}:O{incomes_end_row})").font = bold_font
                ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
                ws.cell(row=row_cursor, column=15).border = border_total
                ws.cell(row=row_cursor, column=15).fill = fill_accent
                incomes_tot_row = row_cursor
                row_cursor += 2
            
            # Sobra / Saldo Líquido
            ws.cell(row=row_cursor, column=1, value="FLUXO DE CAIXA LÍQUIDO (SOBRA)").font = bold_font
            for m in range(1, 13):
                col_letter = get_column_letter(3 + m - 1)
                cell = ws.cell(row=row_cursor, column=3 + m - 1, value=f"={col_letter}{incomes_tot_row}-{col_letter}{expenses_tot_row}")
                cell.number_format = '#,##0.00'
                cell.font = bold_font
                cell.border = border_total
                cell.fill = fill_total
            ws.cell(row=row_cursor, column=15, value=f"=O{incomes_tot_row}-O{expenses_tot_row}").font = bold_font
            ws.cell(row=row_cursor, column=15).number_format = '#,##0.00'
            ws.cell(row=row_cursor, column=15).border = border_total
            ws.cell(row=row_cursor, column=15).fill = fill_total
            row_cursor += 1
            
        # Auto-adjust columns widths
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if cell.row == 1: continue # ignore main title length
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(EXPORT_PATH)
    return jsonify({"success": True, "message": "Planilha exportada com sucesso", "file": EXPORT_PATH})

if __name__ == '__main__':
    # Ensure static directory exists
    os.makedirs(os.path.join(os.path.dirname(__file__), 'static'), exist_ok=True)
    app.run(host='0.0.0.0', port=5001, debug=True)
