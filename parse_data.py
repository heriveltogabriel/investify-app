import openpyxl
import json
import os
import datetime

EXCEL_PATH = '/Users/hgs/Documents/investimento/Investimentos.xlsx'
JSON_PATH = '/Users/hgs/Documents/investimento/db.json'

def parse_double_col_year(ws):
    """
    Parses a sheet with 2 columns per month (like 2024).
    """
    # 2024 has columns:
    # C: Situação Jan 24 (Initial balance)
    # D, E: Jan Aporte, Jan Juros
    # F, G: Feb Aporte, Feb Juros
    # ...
    # Z, AA: Dec Aporte, Dec Juros
    months_names = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho", 
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    
    # Let's map row contents:
    # Row 3: CDB
    # Row 4: IVVB11
    # Row 5: FII
    # Row 6: Fundos
    
    data = {}
    
    # Helper to clean numeric values
    def val(x):
        if x is None: return 0.0
        try: return float(x)
        except: return 0.0

    # Let's find initial values (Dec 23 / Jan 24 starting)
    # For CDB, starting sits in C3
    cdb_init = val(ws.cell(row=3, column=3).value)
    ivvb_init = val(ws.cell(row=4, column=3).value)
    fii_init = val(ws.cell(row=5, column=3).value)
    fundos_init = val(ws.cell(row=6, column=3).value)
    
    # We will accumulate balances month by month
    current_balances = {
        "cdb": cdb_init,
        "ivvb11": ivvb_init,
        "fii": fii_init,
        "fundos": fundos_init
    }
    
    for i, m_name in enumerate(months_names):
        col_aporte = 4 + 2 * i # Col D, F, H, J, ...
        col_juros = 5 + 2 * i  # Col E, G, I, K, ...
        
        # Read inputs for CDB
        cdb_ap = val(ws.cell(row=3, column=col_aporte).value)
        cdb_jr = val(ws.cell(row=3, column=col_juros).value)
        
        # Read inputs for IVVB11
        ivvb_ap = val(ws.cell(row=4, column=col_aporte).value)
        ivvb_jr = val(ws.cell(row=4, column=col_juros).value)
        
        # Read inputs for FII
        fii_ap = val(ws.cell(row=5, column=col_aporte).value)
        fii_jr = val(ws.cell(row=5, column=col_juros).value)
        
        # Read inputs for Fundos
        fundos_ap = val(ws.cell(row=6, column=col_aporte).value)
        fundos_jr = val(ws.cell(row=6, column=col_juros).value)
        
        # Calculate new balances
        # Balance = Previous Balance + Aporte + Juros
        current_balances["cdb"] += cdb_ap + cdb_jr
        current_balances["ivvb11"] += ivvb_ap + ivvb_jr
        current_balances["fii"] += fii_ap + fii_jr
        current_balances["fundos"] += fundos_ap + fundos_jr
        
        # Store in standard schema (group by bank for unity, in 2024 we place under 'outros')
        data[str(i+1)] = {
            "investments": {
                "itau": {
                    "cdb": current_balances["cdb"],
                    "aporte": cdb_ap,
                    "juros": cdb_jr
                },
                "outros": {
                    "ivvb11": { "saldo": current_balances["ivvb11"], "aporte": ivvb_ap, "juros": ivvb_jr },
                    "fii": { "saldo": current_balances["fii"], "aporte": fii_ap, "juros": fii_jr },
                    "fundos": { "saldo": current_balances["fundos"], "aporte": fundos_ap, "juros": fundos_jr }
                }
            },
            "expenses": { "fixed": {}, "cards": {} },
            "incomes": { "salario": 0.0, "aluguel": 0.0, "extra": 0.0 }
        }
        
    return data

def parse_single_col_year(ws, year):
    """
    Parses a sheet with 1 column per month (like 2025, 2026).
    """
    # In 2025:
    # Month M is in column M + 3 (Col D is Jan) for both investments and expenses.
    # In 2026:
    # Month M is in column M + 2 (Col C is Jan) for investments,
    # and column M + 3 (Col D is Jan) for expenses.
    
    import unicodedata

    def normalize(t):
        if not t: return ""
        s = str(t).strip().upper()
        s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
        return s.replace("-", " ")

    def val(x):
        if x is None: return 0.0
        try: return float(x)
        except: return 0.0

    # Default row fallbacks (based on historic positions)
    itau_cdb_row = 4
    itau_jr_row = 5
    itau_ap_row = 6
    itau_total_row = 8
    
    bb_cdb_row = 11
    bb_jr_row = 12
    bb_ap_row = 13
    bb_total_row = 14
    
    c6_cdb_row = 17
    c6_jr_row = 18
    c6_ap_row = 19
    c6_total_row = 20
    
    portfolio_total_row = 24 if year == 2025 else 23

    # Dynamic row detection
    row_mapping = {}
    current_bank = None
    current_sec = None
    
    for r in range(1, 100):
        val_a = normalize(ws.cell(row=r, column=1).value)
        val_c = normalize(ws.cell(row=r, column=3).value)
        
        # 1. Investments section bank headers & total row detection
        if "ITAU" in val_a and "JUROS" not in val_a and "APORTE" not in val_a:
            current_bank = "itau"
        elif "BANCO DO BRASIL" in val_a:
            current_bank = "bb"
        elif "C6 BANK" in val_a:
            current_bank = "c6"
            
        if current_bank == "itau":
            if "CDB" in val_a: itau_cdb_row = r
            elif "JUROS" in val_a: itau_jr_row = r
            elif "APORTE" in val_a: itau_ap_row = r
        elif current_bank == "bb":
            if "CDB" in val_a: bb_cdb_row = r
            elif "JUROS" in val_a: bb_jr_row = r
            elif "APORTE" in val_a: bb_ap_row = r
        elif current_bank == "c6":
            if "CDB" in val_a: c6_cdb_row = r
            elif "JUROS" in val_a: c6_jr_row = r
            elif "APORTE" in val_a: c6_ap_row = r
            
        if val_a == "TOTAL":
            if current_bank == "itau":
                itau_total_row = r
            elif current_bank == "bb":
                bb_total_row = r
            elif current_bank == "c6":
                c6_total_row = r
            current_bank = None
            
        if "TOTAL DO MES" in val_a or "TOTAL DO MÊS" in val_a:
            portfolio_total_row = r

        # 2. Section detection for expenses/cards/incomes
        if "GASTOS FIXOS" in val_a:
            current_sec = "fixed"
        elif "CARTAO" in val_a or "CREDITO" in val_a:
            current_sec = "cards"
        elif "RECEITAS DO MES" in val_a or ("RECEITAS" in val_a and "DESPESAS" not in val_a):
            current_sec = "incomes"
        elif "TOTAL DESPESAS" in val_a or "RECEITAS VS DESPESAS" in val_a:
            current_sec = None
            
        if not current_sec:
            continue
            
        # Label matching within sections (reading from Column C)
        if current_sec == "fixed":
            if "CONDOMINIO" in val_c:
                row_mapping["condominio"] = r
            elif "AGUA" in val_c:
                row_mapping["agua"] = r
            elif "LUZ" in val_c:
                row_mapping["luz"] = r
            elif "NET" in val_c:
                row_mapping["net"] = r
            elif "CELULAR" in val_c:
                row_mapping["celular"] = r
            elif "CARRO" in val_c or "IPVA" in val_c:
                row_mapping["carro_ipva_iptu"] = r
            elif "INGLES" in val_c:
                row_mapping["ingles"] = r
            elif "SEGURO" in val_c:
                row_mapping["itau_seguro"] = r
            elif "FAXINA" in val_c:
                row_mapping["faxina"] = r
                
        elif current_sec == "cards":
            if "MASTERCARD" in val_c or "ITAU" in val_c:
                row_mapping["mastercard_itau"] = r
            elif "VISA" in val_c or "C6" in val_c:
                row_mapping["visa_c6"] = r
            elif "ELO" in val_c or "BB" in val_c:
                row_mapping["elo_bb"] = r
                
        elif current_sec == "incomes":
            if "DIA 15" in val_c:
                row_mapping["salario_dia_15"] = r
            elif "DIA 30" in val_c:
                row_mapping["salario_dia_30"] = r
            elif "SALARIO" in val_c and "DIA" not in val_c and "SOMATORIO" not in val_c:
                row_mapping["salario"] = r
            elif "ALUGUEL" in val_c:
                row_mapping["aluguel"] = r
            elif "EXTRA" in val_c or "FERIAS" in val_c or "PLR" in val_c:
                row_mapping["extra"] = r

    data = {}

    for m in range(1, 13):
        month_num = str(m)
        
        # Determine column indexes based on year and section
        inv_col = m + 3 if year == 2025 else m + 2
        exp_col = m + 3
        
        # Read CDB Balance, Aporte and Juros
        itau_cdb = val(ws.cell(row=itau_cdb_row, column=inv_col).value)
        itau_ap = val(ws.cell(row=itau_ap_row, column=inv_col).value)
        itau_jr = val(ws.cell(row=itau_jr_row, column=inv_col).value)
        
        bb_cdb = val(ws.cell(row=bb_cdb_row, column=inv_col).value)
        bb_ap = val(ws.cell(row=bb_ap_row, column=inv_col).value)
        bb_jr = val(ws.cell(row=bb_jr_row, column=inv_col).value)
        
        c6_cdb = val(ws.cell(row=c6_cdb_row, column=inv_col).value)
        c6_ap = val(ws.cell(row=c6_ap_row, column=inv_col).value)
        c6_jr = val(ws.cell(row=c6_jr_row, column=inv_col).value)
        
        # Fix the BB Jan 25 interest calculation error in the sheet
        if year == 2025 and month_num == "1":
            # BB Jan 25 CDB was 18,745.75 and Dec 24 CDB was 18,529.00
            # Juros should be 18745.75 - 18529.00 = 216.75
            bb_jr = 216.75
            
        # Read totals directly
        itau_tot = val(ws.cell(row=itau_total_row, column=inv_col).value) if itau_total_row else 0.0
        bb_tot = val(ws.cell(row=bb_total_row, column=inv_col).value) if bb_total_row else 0.0
        c6_tot = val(ws.cell(row=c6_total_row, column=inv_col).value) if c6_total_row else 0.0
        portfolio_tot = val(ws.cell(row=portfolio_total_row, column=inv_col).value) if portfolio_total_row else 0.0
            
        # Expenses
        fixed_exp = {}
        fixed_keys = ["condominio", "agua", "luz", "net", "celular", "carro_ipva_iptu", "ingles", "itau_seguro", "faxina"]
        for key in fixed_keys:
            if key in row_mapping:
                fixed_exp[key] = val(ws.cell(row=row_mapping[key], column=exp_col).value)
            
        cards_exp = {}
        card_keys = ["mastercard_itau", "visa_c6", "elo_bb"]
        for key in card_keys:
            if key in row_mapping:
                cards_exp[key] = val(ws.cell(row=row_mapping[key], column=exp_col).value)
            
        # Incomes
        inc = {
            "salario": 0.0,
            "aluguel": 0.0,
            "extra": 0.0
        }
        income_keys = ["salario_dia_15", "salario_dia_30", "salario", "aluguel", "extra"]
        for key in income_keys:
            if key in row_mapping:
                inc[key] = val(ws.cell(row=row_mapping[key], column=exp_col).value)
            
        if year == 2025:
            # Set default income values for 2025 to keep consistency
            inc["salario"] = 17000.0
            inc["aluguel"] = 3650.0
            inc["extra"] = 0.0
            
        data[month_num] = {
            "investments": {
                "itau": { "cdb": itau_cdb, "aporte": itau_ap, "juros": itau_jr, "total": itau_tot },
                "bb": { "cdb": bb_cdb, "aporte": bb_ap, "juros": bb_jr, "total": bb_tot },
                "c6": { "cdb": c6_cdb, "aporte": c6_ap, "juros": c6_jr, "total": c6_tot }
            },
            "total_carteira": portfolio_tot,
            "expenses": {
                "fixed": fixed_exp,
                "cards": cards_exp
            },
            "incomes": inc
        }
        
    return data

def main(filepath=None):
    path = filepath if filepath else EXCEL_PATH
    wb = openpyxl.load_workbook(path, data_only=True)
    db = {}
    
    if "2024" in wb.sheetnames:
        db["2024"] = parse_double_col_year(wb["2024"])
        
    if "2025" in wb.sheetnames:
        db["2025"] = parse_single_col_year(wb["2025"], 2025)
        
    if "2026" in wb.sheetnames:
        db["2026"] = parse_single_col_year(wb["2026"], 2026)
        
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2, ensure_ascii=False)
        
    print(f"Data parsed successfully and written to {JSON_PATH}")

if __name__ == "__main__":
    main()
